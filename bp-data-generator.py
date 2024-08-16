import openpyxl


def generate_ranges(starting_number, range_size, total_numbers):
    ranges = []
    current_number = starting_number
    end_number = starting_number + total_numbers - 1
    while current_number <= end_number:
        range_from = current_number
        range_to = min(current_number + range_size - 1, end_number)
        ranges.append((range_from, range_to))
        current_number = range_to + 1  # Move to the next range
    return ranges

def create_output_file(starting_number, range_size, total_numbers, references, uim, output_file):
    ranges = generate_ranges(starting_number, range_size, total_numbers)
    ref_index = 0
    remaining_counts = {ref: count for ref, count in references}  # Keep track of remaining counts for each reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Cover", "Blank", "rangeFrom", "rangeTo", "ballotNo", "PollingStationRef", "UIM", "Tendered", "bookOf",
               "MDbarcode"])

    for range_from, range_to in ranges:
        reference, num_occurrences = references[ref_index]

        while remaining_counts[reference] == 0:
            # Move to the next reference if the current one is exhausted
            ref_index = (ref_index + 1) % len(references)
            reference, num_occurrences = references[ref_index]

        ws.append(["TRUE", "FALSE", range_from, range_to, "", reference, uim, tendered, "", MDbarcode])
        remaining_counts[reference] -= 1

        for num in range(range_from, range_to + 1):
            ws.append(["FALSE", "FALSE", "", "", num, "", uim, tendered, "", MDbarcode])

        ws.append(["FALSE", "TRUE", "", "", "", "", uim, tendered, "", MDbarcode])

    # Update the "bookOf" column
    cover_row_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == "TRUE":
                cover_row_count += 1
                cell.offset(column=8).value = f"{cover_row_count} of {len(ranges)}"

    wb.save(output_file)


# Example usage:
MDbarcode = "TRUE"
starting_number = 400001
range_size = 10
total_numbers = 800
references = [("1", 5), ("2", 5), ("3", 9), ("4", 6), ("5", 3)]
uim = "NVGA"
tendered = "FALSE"
output_file = "output.xlsx"

create_output_file(starting_number, range_size, total_numbers, references, uim, output_file)
